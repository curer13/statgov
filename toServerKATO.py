import pyodbc
import openpyxl
from os import listdir
from os.path import isfile, join
import datetime
import re

server = 'localhost'
database = 'juridicalFaceDb'
username = 'sa'
password = 'Abz09jvv1'

# Connect to SQL Server database
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' +
                      server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)

# Create cursor
cursor = conn.cursor()


def findLastRequest():
    zayavki = 'zayavki/'
    onlyfiles = {}

    for f in listdir(zayavki):
        filePath=join(zayavki, f)
        if isfile(filePath) and f != '.DS_Store' and not re.search(r'^(~\$)', f):
            onlyfiles[filePath] = None

    print(onlyfiles)

    for file in onlyfiles:
        d_m_y, H_M = getDateFromFile(file)
        req_date = datetime.datetime.strptime(d_m_y + '_' + H_M, '%d-%m-%y_%H-%M')
        onlyfiles[file] = req_date

    return max(onlyfiles, key=lambda req: onlyfiles[req])


def getDateFromFile(file):
    splitlist = file.split('_')
    # print(splitlist)
    d_m_y = splitlist[1]
    H_M = splitlist[2].split('.')[0]
    return d_m_y, H_M


def getShortName(fullname):
    matches = re.findall(r'\"(.*)\"', fullname)
    return matches[0]


def sheetToSQL(sheet, startingRow):
    global cursor, conn

    query = """
    IF NOT EXISTS(select bin from juridicalInfo where bin=?)
    BEGIN
        INSERT INTO  juridicalInfo(
                bin,
                tolyk_atauy,
                polnoe_naimenovanie,
                data_registracii,
                oked,
                negizgi_kyzmet_turinin_atauy,
                naimenovanie_osnovnogo_vida_deyatelnosti,
                vtoroy_oked,
                krp,
                kkzh_atauy,
                naimenovanie_krp,
                kato,
                eldin_mekeninin_atauy,
                naimenovanie_naselennogo_punkta,
                juridicheskiy_adres,
                basshynyn_tae_fio_rukovoditelya
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    END
    """

    # grab existing row count in the database for validation later
    # cursor.execute("SELECT count(*) FROM juridicalInfo")
    # before_import = cursor.fetchone()

    nrows=0

    for r in range(startingRow, sheet.max_row + 1):
        bin = sheet.cell(r, 1).value
        # short_name=getShortName(sheet.cell(r, 3).value)
        tolyk_atauy = sheet.cell(r, 2).value
        polnoe_naimenovanie = sheet.cell(r, 3).value
        data_registracii = sheet.cell(r, 4).value
        oked = sheet.cell(r, 5).value
        negizgi_kyzmet_turinin_atauy = sheet.cell(r, 6).value
        naimenovanie_osnovnogo_vida_deyatelnosti = sheet.cell(r, 7).value
        vtoroy_oked = sheet.cell(r, 8).value
        krp = sheet.cell(r, 9).value
        kkzh_atauy = sheet.cell(r, 10).value
        naimenovanie_krp = sheet.cell(r, 11).value
        kato = sheet.cell(r, 12).value
        eldin_mekeninin_atauy = sheet.cell(r, 13).value
        naimenovanie_naselennogo_punkta = sheet.cell(r, 14).value
        juridicheskiy_adres = sheet.cell(r, 15).value
        basshynyn_tae_fio_rukovoditelya = sheet.cell(r, 16).value

        # Assign values from each row
        values = (bin, bin, tolyk_atauy, polnoe_naimenovanie, data_registracii, oked,
                  negizgi_kyzmet_turinin_atauy, naimenovanie_osnovnogo_vida_deyatelnosti, vtoroy_oked, krp, kkzh_atauy,
                  naimenovanie_krp, kato, eldin_mekeninin_atauy, naimenovanie_naselennogo_punkta, juridicheskiy_adres,
                  basshynyn_tae_fio_rukovoditelya)

        # Execute sql Query
        cursor.execute(query, values)

        # Commit the transaction
        conn.commit()

        nrows+=1
        print(f'{nrows} rows was inserted')

    # If you want to check if all rows are imported
    # cursor.execute("SELECT count(*) FROM juridicalInfo")
    # result = cursor.fetchone()

    # print((result[0] - before_import[0]) == len(data.index))  # should be True


def fromExcelToSQL(excelFileName):
    global cursor, conn

    # with open('request_name.txt', 'r') as file:
    #   excelFileName = file.read()

    # Open the workbook

    workbook = openpyxl.load_workbook(excelFileName)

    print('Workbook was downloaded')

    query1 = """
    IF NOT EXISTS(SELECT [name] FROM sys.tables WHERE [name] = 'accounts')
    BEGIN
        SET DATEFORMAT dmy;
        CREATE TABLE juridicalInfo(
            bin varchar(255),
            tolyk_atauy ntext,
            polnoe_naimenovanie ntext,
            data_registracii date,
            oked varchar(255),
            negizgi_kyzmet_turinin_atauy ntext,
            naimenovanie_osnovnogo_vida_deyatelnosti ntext,
            vtoroy_oked varchar(255),
            krp varchar(25),
            kkzh_atauy ntext,
            naimenovanie_krp ntext,
            kato varchar(255),
            eldin_mekeninin_atauy ntext,
            naimenovanie_naselennogo_punkta ntext,
            juridicheskiy_adres ntext,
            basshynyn_tae_fio_rukovoditelya ntext
        )
    
        CREATE index idx_bin
        ON juridicalInfo(bin)
    END
    """

    try:
        cursor.execute(query1)
        conn.commit()
    except pyodbc.ProgrammingError:
        pass

    try:
        sheetToSQL(workbook[workbook.sheetnames[0]], 3)
    except pyodbc.ProgrammingError:
        pass

    if len(workbook.sheetnames)>1:
        for sheetname in workbook.sheetnames[1:]:
            sheet = workbook[sheetname]
            try:
                sheetToSQL(sheet, 2)
            except pyodbc.ProgrammingError:
                pass

    # Close the database connection
    conn.close()


def main():
    excelFile = findLastRequest()
    # excelFile='testbook.xlsx'
    print(f'Importing {excelFile}')
    fromExcelToSQL(excelFile)


if __name__ == '__main__':
    main()
